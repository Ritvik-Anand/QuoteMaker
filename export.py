import io
import os
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ReportLab's built-in Helvetica/Vera fonts have no glyph for ₹ (U+20B9) — it
# renders as a solid black box. DejaVu Sans covers it, so we embed it instead.
_FONTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
pdfmetrics.registerFont(TTFont("DejaVuSans", os.path.join(_FONTS_DIR, "DejaVuSans.ttf")))
pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", os.path.join(_FONTS_DIR, "DejaVuSans-Bold.ttf")))
FONT = "DejaVuSans"
FONT_BOLD = "DejaVuSans-Bold"

BRAND_COLOR = colors.HexColor("#2c5f8a")
ACCENT_COLOR = colors.HexColor("#eef4fa")
LIGHT_GRAY = colors.HexColor("#f7f7f7")
MID_GRAY = colors.HexColor("#dddddd")


def _compute_totals(quote_items, gst_rate, cash_discount=False):
    subtotal = sum(item["quantity"] * item["final_price"] for item in quote_items)
    # Cash discount comes off the subtotal before GST is applied, not off
    # the final total — GST is owed on what the client actually pays.
    discount = round(subtotal * 0.01, 2) if cash_discount else 0
    taxable = subtotal - discount
    gst_amount = round(taxable * gst_rate / 100, 2)
    total = round(taxable + gst_amount, 2)
    return round(subtotal, 2), discount, gst_amount, total


def generate_pdf(quotation: dict, quote_items: list[dict]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    story = []

    # Header
    header_data = [[
        Paragraph(
            f'<font color="#2c5f8a" size="20"><b>BAGULA MUKHI</b></font><br/>'
            f'<font color="#777777" size="9">Electrical Goods Supplier</font>',
            ParagraphStyle("h", fontName=FONT, alignment=TA_LEFT)
        ),
        Paragraph(
            f'<font color="#2c5f8a" size="14"><b>QUOTATION</b></font><br/>'
            f'<font color="#555555" size="9">No: <b>{quotation["quote_number"]}</b></font><br/>'
            f'<font color="#555555" size="9">Date: {quotation["date"]}</font>',
            ParagraphStyle("h2", fontName=FONT, alignment=TA_RIGHT)
        ),
    ]]
    header_table = Table(header_data, colWidths=[95 * mm, 85 * mm])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(header_table)
    story.append(HRFlowable(width="100%", thickness=1, color=MID_GRAY, spaceAfter=6))

    # Client block
    client_info = f"<b>To:</b> {quotation['client_name']}"
    if quotation.get("client_address"):
        client_info += f"<br/>{quotation['client_address'].replace(chr(10), '<br/>')}"
    story.append(Paragraph(client_info, ParagraphStyle(
        "client", fontName=FONT, fontSize=10, leading=14, spaceAfter=8, textColor=colors.HexColor("#333333")
    )))

    # Items table
    col_widths = [12 * mm, 20 * mm, 68 * mm, 15 * mm, 15 * mm, 22 * mm, 25 * mm]
    table_data = [[
        Paragraph("<b>#</b>", ParagraphStyle("th", fontName=FONT_BOLD, fontSize=8, alignment=TA_CENTER)),
        Paragraph("<b>Code</b>", ParagraphStyle("th", fontName=FONT_BOLD, fontSize=8, alignment=TA_CENTER)),
        Paragraph("<b>Description</b>", ParagraphStyle("th", fontName=FONT_BOLD, fontSize=8)),
        Paragraph("<b>Unit</b>", ParagraphStyle("th", fontName=FONT_BOLD, fontSize=8, alignment=TA_CENTER)),
        Paragraph("<b>Qty</b>", ParagraphStyle("th", fontName=FONT_BOLD, fontSize=8, alignment=TA_CENTER)),
        Paragraph("<b>Rate (₹)</b>", ParagraphStyle("th", fontName=FONT_BOLD, fontSize=8, alignment=TA_RIGHT)),
        Paragraph("<b>Amount (₹)</b>", ParagraphStyle("th", fontName=FONT_BOLD, fontSize=8, alignment=TA_RIGHT)),
    ]]

    for i, item in enumerate(quote_items, 1):
        amount = item["quantity"] * item["final_price"]
        row_style = ParagraphStyle("td", fontName=FONT, fontSize=8, leading=11)
        row_style_r = ParagraphStyle("tdr", fontName=FONT, fontSize=8, leading=11, alignment=TA_RIGHT)
        row_style_c = ParagraphStyle("tdc", fontName=FONT, fontSize=8, leading=11, alignment=TA_CENTER)
        table_data.append([
            Paragraph(str(i), row_style_c),
            Paragraph(item.get("code") or "-", row_style_c),
            Paragraph(item["description"], row_style),
            Paragraph(item.get("unit", "Nos"), row_style_c),
            Paragraph(_fmt_num(item["quantity"]), row_style_c),
            Paragraph(f"{item['final_price']:,.2f}", row_style_r),
            Paragraph(f"{amount:,.2f}", row_style_r),
        ])

    items_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    ts = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ACCENT_COLOR),
        ("TEXTCOLOR", (0, 0), (-1, 0), BRAND_COLOR),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.5, MID_GRAY),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ])
    items_table.setStyle(ts)
    story.append(items_table)
    story.append(Spacer(1, 4 * mm))

    # Totals
    cash_discount = bool(quotation.get("cash_discount"))
    subtotal, discount, gst_amount, total = _compute_totals(
        quote_items, quotation.get("gst_rate", 18), cash_discount
    )
    gst_rate = quotation.get("gst_rate", 18)

    totals_data = [["", "Subtotal", f"₹ {subtotal:,.2f}"]]
    if cash_discount:
        totals_data.append(["", "Cash Discount (1%)", f"− ₹ {discount:,.2f}"])
    totals_data.append(["", f"GST ({gst_rate:.0f}%)", f"₹ {gst_amount:,.2f}"])
    totals_data.append(["", "TOTAL", f"₹ {total:,.2f}"])
    total_row = len(totals_data) - 1
    totals_table = Table(totals_data, colWidths=[115 * mm, 35 * mm, 27 * mm])
    totals_table.setStyle(TableStyle([
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (-1, -1), FONT),
        ("FONTNAME", (1, total_row), (-1, total_row), FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTSIZE", (1, total_row), (-1, total_row), 10),
        ("LINEABOVE", (1, total_row), (-1, total_row), 1, BRAND_COLOR),
        ("TEXTCOLOR", (1, total_row), (-1, total_row), BRAND_COLOR),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(totals_table)

    # Notes
    if quotation.get("notes"):
        story.append(Spacer(1, 4 * mm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=MID_GRAY, spaceAfter=4))
        story.append(Paragraph(
            f"<b>Notes:</b> {quotation['notes']}",
            ParagraphStyle("notes", fontName=FONT, fontSize=8, textColor=colors.HexColor("#555555"))
        ))

    # Footer
    story.append(Spacer(1, 6 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MID_GRAY, spaceAfter=3))
    story.append(Paragraph(
        "Thank you for your business. This is a computer-generated quotation.",
        ParagraphStyle("footer", fontName=FONT, fontSize=7, textColor=colors.HexColor("#999999"), alignment=TA_CENTER)
    ))

    doc.build(story)
    return buffer.getvalue()


def generate_excel(quotation: dict, quote_items: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    brand = "1a3c5e"
    accent = "e8f0fe"
    light = "f5f5f5"

    thin = Side(style="thin", color="cccccc")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Col widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16

    row = 1

    # Business name
    ws.merge_cells(f"A{row}:D{row}")
    c = ws[f"A{row}"]
    c.value = "BAGULA MUKHI"
    c.font = Font(name="Calibri", size=18, bold=True, color=brand)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28

    ws.merge_cells(f"E{row}:G{row}")
    c = ws[f"E{row}"]
    c.value = "QUOTATION"
    c.font = Font(name="Calibri", size=16, bold=True, color=brand)
    c.alignment = Alignment(horizontal="right", vertical="center")
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"].value = "Electrical Goods Supplier"
    ws[f"A{row}"].font = Font(name="Calibri", size=9, color="555555")

    ws.merge_cells(f"E{row}:G{row}")
    c = ws[f"E{row}"]
    c.value = f"No: {quotation['quote_number']}    Date: {quotation['date']}"
    c.font = Font(name="Calibri", size=9)
    c.alignment = Alignment(horizontal="right")
    row += 2

    # Client
    ws[f"A{row}"].value = "To:"
    ws[f"A{row}"].font = Font(bold=True, size=10)
    ws.merge_cells(f"B{row}:G{row}")
    ws[f"B{row}"].value = quotation["client_name"]
    ws[f"B{row}"].font = Font(size=10, bold=True)
    row += 1

    if quotation.get("client_address"):
        ws.merge_cells(f"B{row}:G{row}")
        ws[f"B{row}"].value = quotation["client_address"]
        ws[f"B{row}"].font = Font(size=9, color="444444")
        row += 1

    row += 1

    # Header row
    headers = ["#", "Code", "Description", "Unit", "Qty", "Rate (₹)", "Amount (₹)"]
    cols = ["A", "B", "C", "D", "E", "F", "G"]
    for col, h in zip(cols, headers):
        c = ws[f"{col}{row}"]
        c.value = h
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=brand)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[row].height = 18
    row += 1

    # Items
    for i, item in enumerate(quote_items, 1):
        amount = item["quantity"] * item["final_price"]
        fill = PatternFill("solid", fgColor="FFFFFF") if i % 2 == 1 else PatternFill("solid", fgColor=light)
        values = [i, item.get("code") or "", item["description"], item.get("unit", "Nos"),
                  item["quantity"], item["final_price"], amount]
        aligns = ["center", "center", "left", "center", "center", "right", "right"]
        for col, val, align in zip(cols, values, aligns):
            c = ws[f"{col}{row}"]
            c.value = val
            c.font = Font(name="Calibri", size=9)
            c.fill = fill
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = border
            if col in ("F", "G"):
                c.number_format = '#,##0.00'
        row += 1

    row += 1

    # Totals
    cash_discount = bool(quotation.get("cash_discount"))
    subtotal, discount, gst_amount, total = _compute_totals(
        quote_items, quotation.get("gst_rate", 18), cash_discount
    )
    gst_rate = quotation.get("gst_rate", 18)

    totals_lines = [("Subtotal", subtotal)]
    if cash_discount:
        totals_lines.append(("Cash Discount (1%)", -discount))
    totals_lines.append((f"GST ({gst_rate:.0f}%)", gst_amount))

    for label, value in totals_lines:
        ws.merge_cells(f"A{row}:F{row}")
        c = ws[f"A{row}"]
        c.value = label
        c.font = Font(name="Calibri", size=9)
        c.alignment = Alignment(horizontal="right")
        c = ws[f"G{row}"]
        c.value = value
        c.font = Font(name="Calibri", size=9)
        c.alignment = Alignment(horizontal="right")
        c.number_format = '#,##0.00'
        row += 1

    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = "TOTAL"
    c.font = Font(name="Calibri", size=11, bold=True, color=brand)
    c.alignment = Alignment(horizontal="right")
    c.fill = PatternFill("solid", fgColor=accent)
    c = ws[f"G{row}"]
    c.value = total
    c.font = Font(name="Calibri", size=11, bold=True, color=brand)
    c.alignment = Alignment(horizontal="right")
    c.fill = PatternFill("solid", fgColor=accent)
    c.number_format = '#,##0.00'
    row += 2

    # Notes
    if quotation.get("notes"):
        ws[f"A{row}"].value = "Notes:"
        ws[f"A{row}"].font = Font(bold=True, size=9)
        ws.merge_cells(f"B{row}:G{row}")
        ws[f"B{row}"].value = quotation["notes"]
        ws[f"B{row}"].font = Font(size=9, color="555555")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


_MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]

_PAYROLL_STATUS_LABEL = {"draft": "Draft", "finalized": "Sent to Accountant", "paid": "Paid"}


def generate_payroll_pdf(year: int, month: int, rows: list[dict]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    story = []

    header_data = [[
        Paragraph(
            f'<font color="#2c5f8a" size="20"><b>BAGULA MUKHI</b></font><br/>'
            f'<font color="#777777" size="9">Electrical Goods Supplier</font>',
            ParagraphStyle("h", fontName=FONT, alignment=TA_LEFT)
        ),
        Paragraph(
            f'<font color="#2c5f8a" size="14"><b>PAYROLL SUMMARY</b></font><br/>'
            f'<font color="#555555" size="9">{_MONTH_NAMES[month]} {year}</font>',
            ParagraphStyle("h2", fontName=FONT, alignment=TA_RIGHT)
        ),
    ]]
    header_table = Table(header_data, colWidths=[95 * mm, 85 * mm])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(header_table)
    story.append(HRFlowable(width="100%", thickness=1, color=MID_GRAY, spaceAfter=6))
    story.append(Spacer(1, 4 * mm))

    col_widths = [42 * mm, 25 * mm, 18 * mm, 18 * mm, 18 * mm, 27 * mm, 27 * mm, 25 * mm]
    th = lambda t, align=TA_LEFT: Paragraph(f"<b>{t}</b>", ParagraphStyle("th", fontName=FONT_BOLD, fontSize=8, alignment=align))
    table_data = [[
        th("Employee"), th("Salary (₹)", TA_RIGHT), th("Present", TA_CENTER), th("Absent", TA_CENTER),
        th("Unmarked", TA_CENTER), th("Computed Pay (₹)", TA_RIGHT), th("Final Pay (₹)", TA_RIGHT), th("Status"),
    ]]

    total_computed = 0.0
    total_final = 0.0
    for r in rows:
        unmarked = r["calendar_days"] - r["present_days"] - r["absent_days"]
        total_computed += r["computed_pay"]
        total_final += r["final_pay"]
        row_style = ParagraphStyle("td", fontName=FONT, fontSize=8, leading=11)
        row_style_r = ParagraphStyle("tdr", fontName=FONT, fontSize=8, leading=11, alignment=TA_RIGHT)
        row_style_c = ParagraphStyle("tdc", fontName=FONT, fontSize=8, leading=11, alignment=TA_CENTER)
        table_data.append([
            Paragraph(r["employee_name"], row_style),
            Paragraph(f"{r['monthly_salary']:,.2f}", row_style_r),
            Paragraph(str(r["present_days"]), row_style_c),
            Paragraph(str(r["absent_days"]), row_style_c),
            Paragraph(str(unmarked), row_style_c),
            Paragraph(f"{r['computed_pay']:,.2f}", row_style_r),
            Paragraph(f"{r['final_pay']:,.2f}", row_style_r),
            Paragraph(_PAYROLL_STATUS_LABEL.get(r["status"], r["status"]), row_style),
        ])

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ACCENT_COLOR),
        ("TEXTCOLOR", (0, 0), (-1, 0), BRAND_COLOR),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.5, MID_GRAY),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)
    story.append(Spacer(1, 4 * mm))

    totals_data = [
        ["", "Total Computed Pay", f"₹ {total_computed:,.2f}"],
        ["", "Total Final Pay", f"₹ {total_final:,.2f}"],
    ]
    totals_table = Table(totals_data, colWidths=[115 * mm, 40 * mm, 45 * mm])
    totals_table.setStyle(TableStyle([
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (-1, -1), FONT),
        ("FONTNAME", (1, -1), (-1, -1), FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTSIZE", (1, -1), (-1, -1), 10),
        ("LINEABOVE", (1, -1), (-1, -1), 1, BRAND_COLOR),
        ("TEXTCOLOR", (1, -1), (-1, -1), BRAND_COLOR),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(totals_table)

    story.append(Spacer(1, 6 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MID_GRAY, spaceAfter=3))
    story.append(Paragraph(
        f"Generated on {datetime.now().strftime('%d %b %Y')}. This is a computer-generated payroll summary.",
        ParagraphStyle("footer", fontName=FONT, fontSize=7, textColor=colors.HexColor("#999999"), alignment=TA_CENTER)
    ))

    doc.build(story)
    return buffer.getvalue()


def _fmt_num(n):
    if n == int(n):
        return str(int(n))
    return str(n)
